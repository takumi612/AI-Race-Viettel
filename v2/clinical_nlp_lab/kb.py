from __future__ import annotations

import gzip
import hashlib
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping

from openpyxl import load_workbook

from .text import normalize_alias
from .kb_contract import (
    KBContractError,
    OrganizerCandidateOccurrence,
    canonical_icd_id,
    normalized_surface_matches,
    validate_icd_record_identity,
)


RXNCONSO_FIELDS = [
    "RXCUI", "LAT", "TS", "LUI", "STT", "SUI", "ISPREF", "RXAUI", "SAUI",
    "SCUI", "SDUI", "SAB", "TTY", "CODE", "STR", "SRL", "SUPPRESS", "CVF"
]

RXNREL_FIELDS = [
    "RXCUI1", "AUI1", "STYPE1", "REL", "RXCUI2", "AUI2", "STYPE2", "RELA",
    "RUI", "SRUI", "SAB", "SL", "DIR", "RG", "SUPPRESS", "CVF"
]


def sha256_file(path: str | Path, block_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        while block := stream.read(block_size):
            digest.update(block)
    return digest.hexdigest()


def _json_bytes(payload: Any) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n").encode("utf-8")


def write_jsonl_gz(path: str | Path, records: Iterable[dict[str, Any]]) -> tuple[int, str]:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with destination.open("wb") as raw_stream:
        with gzip.GzipFile(filename="", mode="wb", fileobj=raw_stream, mtime=0) as compressed:
            for record in records:
                compressed.write(_json_bytes(record))
                count += 1
    return count, sha256_file(destination)


def iter_jsonl_gz(path: str | Path) -> Iterator[dict[str, Any]]:
    with gzip.open(Path(path), "rt", encoding="utf-8", newline="") as stream:
        for line_number, line in enumerate(stream, start=1):
            if not line.strip():
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError as exc:
                raise ValueError(f"Invalid JSONL at {path}:{line_number}: {exc}") from exc


def write_metadata(path: str | Path, payload: dict[str, Any]) -> Path:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(payload, stream, ensure_ascii=False, indent=2, sort_keys=True)
        stream.write("\n")
    return destination


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _canonical_icd10(code: str) -> tuple[str, list[str]]:
    cleaned = code.strip()
    canonical = canonical_icd_id(cleaned)
    suffix = cleaned[len(canonical) :]
    markers = list(dict.fromkeys(suffix))
    return canonical, markers


def build_icd10_dictionary(
    workbook_path: str | Path,
    output_path: str | Path,
    metadata_path: str | Path,
    sheet_name: str = "ICD10",
    header_row: int = 3,
    *,
    source_label: str | None = None,
    artifact_label: str | None = None,
) -> dict[str, Any]:
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing ICD-10 sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [re.sub(r"\s+", " ", _cell_text(cell.value)).upper() for cell in sheet[header_row]]
    positions = {header: index for index, header in enumerate(headers) if header}
    required = ["MÃ BỆNH", "MÃ BỆNH KHÔNG DẤU", "DISEASE NAME", "TÊN BỆNH"]
    missing = [header for header in required if header not in positions]
    if missing:
        raise ValueError(f"Missing ICD-10 headers: {missing}")

    aggregated: dict[str, dict[str, Any]] = {}
    raw_rows = 0
    marker_counts = Counter()
    duplicate_display_rows = 0
    seen_display: set[tuple[str, str, str]] = set()
    missing_english = 0

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        display_code = _cell_text(row[positions["MÃ BỆNH"]])
        code_without_dot = _cell_text(row[positions["MÃ BỆNH KHÔNG DẤU"]])
        name_en = _cell_text(row[positions["DISEASE NAME"]])
        name_vi = _cell_text(row[positions["TÊN BỆNH"]])
        if not any((display_code, code_without_dot, name_en, name_vi)):
            continue
        raw_rows += 1
        canonical_code, markers = _canonical_icd10(display_code)
        if not canonical_code:
            continue
        expected_no_dot = canonical_code.replace(".", "")
        if code_without_dot and expected_no_dot != code_without_dot:
            raise ValueError(
                f"ICD-10 canonical mismatch: {display_code!r} -> {expected_no_dot!r}, sheet has {code_without_dot!r}"
            )
        for marker in markers:
            marker_counts[marker] += 1
        if not name_en:
            missing_english += 1
        display_key = (display_code, name_en, name_vi)
        if display_key in seen_display:
            duplicate_display_rows += 1
        seen_display.add(display_key)

        record = aggregated.setdefault(
            canonical_code,
            {
                "candidate_id": canonical_code,
                "canonical_id": canonical_code,
                "code_with_dot": canonical_code,
                "code_without_dot": expected_no_dot,
                "name_vi": "",
                "name_en": "",
                "aliases": [],
                "detection_aliases": [],
                "display_codes": [],
                "official_display_ids": [],
                "markers": [],
                "source": "ICD10.xlsx",
            },
        )
        if name_vi and not record["name_vi"]:
            record["name_vi"] = name_vi
        if name_en and not record["name_en"]:
            record["name_en"] = name_en
        record["display_codes"].append(display_code)
        record["official_display_ids"].append(display_code)
        record["markers"].extend(markers)
        record["aliases"].extend(value for value in (canonical_code, expected_no_dot, name_en, name_vi) if value)
        record["detection_aliases"].extend(value for value in (name_en, name_vi) if value)

    records: list[dict[str, Any]] = []
    normalized_aliases: set[str] = set()
    for code in sorted(aggregated):
        record = aggregated[code]
        record["aliases"] = list(dict.fromkeys(record["aliases"]))
        record["detection_aliases"] = list(dict.fromkeys(record["detection_aliases"]))
        record["display_codes"] = list(dict.fromkeys(record["display_codes"]))
        record["official_display_ids"] = list(dict.fromkeys(record["official_display_ids"]))
        record["markers"] = list(dict.fromkeys(record["markers"]))
        validate_icd_record_identity(record)
        normalized_aliases.update(normalize_alias(alias) for alias in record["aliases"] if normalize_alias(alias))
        records.append(record)

    written, output_sha = write_jsonl_gz(output_path, records)
    metadata = {
        "artifact": artifact_label or str(Path(output_path).as_posix()),
        "source": source_label or str(workbook_path.as_posix()),
        "source_sha256": sha256_file(workbook_path),
        "sheet": sheet_name,
        "header_row": header_row,
        "raw_row_count": raw_rows,
        "candidate_count": written,
        "normalized_alias_count": len(normalized_aliases),
        "duplicate_display_rows_removed": duplicate_display_rows,
        "missing_english_name_rows": missing_english,
        "marker_counts": {"dagger": marker_counts["†"], "asterisk": marker_counts["*"]},
        "identity_pair_count": sum(len(record["official_display_ids"]) for record in records),
        "multi_display_canonical_count": sum(len(record["official_display_ids"]) > 1 for record in records),
        "schema_version": 2,
        "sha256": output_sha,
    }
    write_metadata(metadata_path, metadata)
    return metadata


def _split_rrf_line(line: str, expected_fields: int) -> list[str]:
    parts = line.rstrip("\r\n").split("|")
    if parts and parts[-1] == "":
        parts.pop()
    if len(parts) != expected_fields:
        raise ValueError(f"Expected {expected_fields} RRF fields, found {len(parts)}")
    return parts


def build_rxnorm_dictionary(
    zip_path: str | Path,
    member: str,
    output_path: str | Path,
    metadata_path: str | Path,
    languages: Iterable[str],
    sources: Iterable[str],
    tty_values: Iterable[str],
    suppress_values: Iterable[str],
    *,
    organizer_requirements: Mapping[
        str, OrganizerCandidateOccurrence | Iterable[OrganizerCandidateOccurrence]
    ]
    | None = None,
    dataset_pair_fingerprint: str | None = None,
    evidence_path: str | Path | None = None,
    evidence_metadata_path: str | Path | None = None,
    expected_supplement_count: int | None = None,
    source_label: str | None = None,
    artifact_label: str | None = None,
    evidence_label: str | None = None,
) -> dict[str, Any]:
    zip_path = Path(zip_path)
    source_zip_sha256 = sha256_file(zip_path)
    languages_set = set(languages)
    sources_set = set(sources)
    tty_set = set(tty_values)
    suppress_set = set(suppress_values)
    if organizer_requirements and not dataset_pair_fingerprint:
        raise KBContractError("A dataset-pair fingerprint is required for organizer supplementation")
    if organizer_requirements and evidence_path is None:
        raise KBContractError("An evidence artifact is required for organizer supplementation")
    row_count = 0
    filtered_count = 0
    tty_counts = Counter()
    candidates: dict[str, dict[str, Any]] = {}
    eligible_supplements: dict[str, list[tuple[int, dict[str, str]]]] = defaultdict(list)
    member_digest = hashlib.sha256()

    with zipfile.ZipFile(zip_path) as archive:
        if member not in archive.namelist():
            raise ValueError(f"Missing RxNorm member: {member}")
        with archive.open(member) as raw_stream:
            for row_count, raw_line in enumerate(raw_stream, start=1):
                member_digest.update(raw_line)
                line = raw_line.decode("utf-8", errors="strict")
                record = dict(zip(RXNCONSO_FIELDS, _split_rrf_line(line, len(RXNCONSO_FIELDS))))
                base_filters_match = (
                    record["LAT"] in languages_set
                    and record["SAB"] in sources_set
                    and record["TTY"] in tty_set
                )
                if (
                    organizer_requirements
                    and record["RXCUI"] in organizer_requirements
                    and base_filters_match
                    and record["SUPPRESS"] in {"O", "E"}
                ):
                    eligible_supplements[record["RXCUI"]].append((row_count, record))
                if not base_filters_match or record["SUPPRESS"] not in suppress_set:
                    continue
                filtered_count += 1
                tty_counts[record["TTY"]] += 1
                candidate = candidates.setdefault(
                    record["RXCUI"],
                    {
                        "candidate_id": record["RXCUI"],
                        "canonical_name": record["STR"],
                        "aliases": [],
                        "detection_aliases": [],
                        "tty": [],
                        "sources": [],
                        "ingredient_ids": [],
                        "brand_names": [],
                        "source": "RxNorm",
                    },
                )
                candidate["aliases"].append(record["STR"])
                candidate["detection_aliases"].append(record["STR"])
                candidate["tty"].append(record["TTY"])
                candidate["sources"].append(record["SAB"])
                if record["TTY"] == "BN":
                    candidate["brand_names"].append(record["STR"])

    requirement_ids = set(organizer_requirements or {})
    missing_ids = requirement_ids - set(candidates)
    supplemental_evidence: list[dict[str, Any]] = []
    supplement_tty_counts: Counter[str] = Counter()
    supplement_suppress_counts: Counter[str] = Counter()
    unexpected_eligible = set(eligible_supplements) - missing_ids
    if unexpected_eligible:
        # O/E rows for a normally admitted ID are irrelevant and must never be selected.
        for candidate_id in unexpected_eligible:
            eligible_supplements.pop(candidate_id, None)
    for candidate_id in sorted(missing_ids, key=lambda value: (len(value), value)):
        matches = eligible_supplements.get(candidate_id, [])
        if len(matches) != 1:
            raise KBContractError(
                f"Expected exactly one eligible O/E row for organizer RxCUI {candidate_id}, found {len(matches)}"
            )
        raw_occurrences = organizer_requirements[candidate_id]  # type: ignore[index]
        if isinstance(raw_occurrences, OrganizerCandidateOccurrence):
            occurrence_rows = (raw_occurrences,)
        else:
            occurrence_rows = tuple(raw_occurrences)
        if len(occurrence_rows) != 1:
            raise KBContractError(
                f"Expected exactly one organizer occurrence for supplemental RxCUI {candidate_id}, "
                f"found {len(occurrence_rows)}"
            )
        occurrence = occurrence_rows[0]
        member_line, record = matches[0]
        if not normalized_surface_matches(occurrence.mention_text, record["STR"]):
            raise KBContractError(
                f"Organizer surface does not match authoritative STR for RxCUI {candidate_id}"
            )
        candidate = {
            "candidate_id": candidate_id,
            "canonical_name": record["STR"],
            "aliases": [record["STR"]],
            "detection_aliases": [],
            "tty": [record["TTY"]],
            "sources": [record["SAB"]],
            "ingredient_ids": [],
            "brand_names": [record["STR"]] if record["TTY"] == "BN" else [],
            "source": "RxNorm",
            "suppress_status": record["SUPPRESS"],
        }
        candidates[candidate_id] = candidate
        supplement_tty_counts[record["TTY"]] += 1
        supplement_suppress_counts[record["SUPPRESS"]] += 1
        supplemental_evidence.append(
            {
                "dataset_pair_fingerprint": dataset_pair_fingerprint,
                "document_id": occurrence.document_id,
                "entity_index": occurrence.entity_index,
                "candidate_id": candidate_id,
                "mention_sha256": occurrence.mention_sha256,
                "source_zip_sha256": source_zip_sha256,
                "member": member,
                "member_sha256": member_digest.hexdigest(),
                "member_line": member_line,
                **{field: record[field] for field in ("RXCUI", "RXAUI", "LAT", "SAB", "TTY", "SUPPRESS", "CODE", "STR")},
                "selection_reason": "organizer_required_suppressed_term",
            }
        )
    if set(eligible_supplements) != missing_ids:
        unresolved = sorted(missing_ids - set(eligible_supplements), key=lambda value: (len(value), value))
        raise KBContractError(f"Unresolved organizer RxNorm IDs: {unresolved[:10]}")
    if expected_supplement_count is not None and len(supplemental_evidence) != expected_supplement_count:
        raise KBContractError(
            f"Expected {expected_supplement_count} organizer supplements, found {len(supplemental_evidence)}"
        )

    records: list[dict[str, Any]] = []
    normalized_aliases: set[str] = set()
    for candidate_id in sorted(candidates, key=lambda value: (len(value), value)):
        candidate = candidates[candidate_id]
        candidate["aliases"] = list(dict.fromkeys(candidate["aliases"]))
        candidate["detection_aliases"] = list(dict.fromkeys(candidate["detection_aliases"]))
        candidate["tty"] = list(dict.fromkeys(candidate["tty"]))
        candidate["sources"] = list(dict.fromkeys(candidate["sources"]))
        candidate["brand_names"] = list(dict.fromkeys(candidate["brand_names"]))
        normalized_aliases.update(normalize_alias(alias) for alias in candidate["aliases"] if normalize_alias(alias))
        records.append(candidate)

    written, output_sha = write_jsonl_gz(output_path, records)
    evidence_sha: str | None = None
    if evidence_path is not None:
        supplemental_evidence.sort(key=lambda row: (len(str(row["candidate_id"])), str(row["candidate_id"])))
        evidence_count, evidence_sha = write_jsonl_gz(evidence_path, supplemental_evidence)
        if evidence_count != len(supplemental_evidence):
            raise KBContractError("Supplement evidence write count mismatch")
        if evidence_metadata_path is not None:
            write_metadata(
                evidence_metadata_path,
                {
                    "artifact": evidence_label or str(Path(evidence_path).as_posix()),
                    "candidate_count": evidence_count,
                    "dataset_pair_fingerprint": dataset_pair_fingerprint,
                    "member": member,
                    "member_sha256": member_digest.hexdigest(),
                    "schema_version": 1,
                    "sha256": evidence_sha,
                    "source": source_label or str(zip_path.as_posix()),
                    "source_sha256": source_zip_sha256,
                },
            )
    metadata = {
        "artifact": artifact_label or str(Path(output_path).as_posix()),
        "source": source_label or str(zip_path.as_posix()),
        "source_sha256": source_zip_sha256,
        "member": member,
        "member_sha256": member_digest.hexdigest(),
        "member_rows_scanned": row_count,
        "filtered_rows": filtered_count,
        "candidate_count": written,
        "normalized_alias_count": len(normalized_aliases),
        "filters": {
            "LAT": sorted(languages_set),
            "SAB": sorted(sources_set),
            "TTY": sorted(tty_set),
            "SUPPRESS": sorted(suppress_set),
        },
        "tty_counts": dict(sorted(tty_counts.items())),
        "normal_candidate_count": written - len(supplemental_evidence),
        "supplement_candidate_count": len(supplemental_evidence),
        "supplement_tty_counts": dict(sorted(supplement_tty_counts.items())),
        "supplement_suppress_counts": dict(sorted(supplement_suppress_counts.items())),
        "supplement_evidence_artifact": evidence_label if evidence_sha else None,
        "supplement_evidence_sha256": evidence_sha,
        "dataset_pair_fingerprint": dataset_pair_fingerprint,
        "schema_version": 2,
        "sha256": output_sha,
    }
    write_metadata(metadata_path, metadata)
    return metadata


def build_rxnorm_relation_cache(
    zip_path: str | Path,
    member: str,
    output_path: str | Path,
    metadata_path: str | Path,
    relation_names: Iterable[str],
    allowed_sources: Iterable[str] = ("RXNORM",),
) -> dict[str, Any]:
    zip_path = Path(zip_path)
    relation_set = set(relation_names)
    source_set = set(allowed_sources)
    scanned = 0
    selected = 0
    missing_rxcui = 0
    relation_counts = Counter()

    def selected_records() -> Iterator[dict[str, Any]]:
        nonlocal scanned, selected, missing_rxcui
        with zipfile.ZipFile(zip_path) as archive:
            if member not in archive.namelist():
                raise ValueError(f"Missing RxNorm relation member: {member}")
            with archive.open(member) as raw_stream:
                with io.TextIOWrapper(raw_stream, encoding="utf-8", errors="strict", newline="") as text_stream:
                    for scanned, line in enumerate(text_stream, start=1):
                        record = dict(zip(RXNREL_FIELDS, _split_rrf_line(line, len(RXNREL_FIELDS))))
                        if record["RELA"] not in relation_set or record["SAB"] not in source_set:
                            continue
                        if not record["RXCUI1"] or not record["RXCUI2"]:
                            missing_rxcui += 1
                            continue
                        selected += 1
                        relation_counts[record["RELA"]] += 1
                        yield {
                            "source_rxcui": record["RXCUI1"],
                            "relation": record["RELA"],
                            "target_rxcui": record["RXCUI2"],
                            "sab": record["SAB"],
                        }

    written, output_sha = write_jsonl_gz(output_path, selected_records())
    metadata = {
        "artifact": str(Path(output_path).as_posix()),
        "source": str(zip_path.as_posix()),
        "source_sha256": sha256_file(zip_path),
        "member": member,
        "member_rows_scanned": scanned,
        "selected_rows": written,
        "skipped_missing_rxcui": missing_rxcui,
        "relations": sorted(relation_set),
        "allowed_sources": sorted(source_set),
        "relation_counts": dict(sorted(relation_counts.items())),
        "sha256": output_sha,
    }
    write_metadata(metadata_path, metadata)
    return metadata


def load_candidate_dictionary(path: str | Path) -> list[dict[str, Any]]:
    return list(iter_jsonl_gz(path))


def verify_dictionary(path: str | Path, expected_count: int | None = None) -> dict[str, Any]:
    count = 0
    ids: set[str] = set()
    duplicate_ids: list[str] = []
    for record in iter_jsonl_gz(path):
        count += 1
        candidate_id = str(record.get("candidate_id", ""))
        if not candidate_id:
            raise ValueError(f"Missing candidate_id in {path} at record {count}")
        if candidate_id in ids:
            duplicate_ids.append(candidate_id)
        ids.add(candidate_id)
    if expected_count is not None and count != expected_count:
        raise ValueError(f"Expected {expected_count} records in {path}, found {count}")
    if duplicate_ids:
        raise ValueError(f"Duplicate candidate IDs in {path}: {duplicate_ids[:10]}")
    return {"path": str(path), "record_count": count, "unique_ids": len(ids), "sha256": sha256_file(path)}
