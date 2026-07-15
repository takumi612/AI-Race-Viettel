import os
import sys
import sqlite3
import json
import openpyxl

# Thêm project root vào sys.path để hỗ trợ import src
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.append(PROJECT_ROOT)

from src.utils.paths import EXCEL_PATH, DB_PATH, JSON_PATH, ICD10_TXT_PATH, RXNORM_TXT_PATH

def setup_database():
    if not os.path.exists(EXCEL_PATH):
        print(f"[ERROR] Excel file not found at: {EXCEL_PATH}")
        print("Please place the official 'ICD10.xlsx' file in the 'data/kb/' directory first.")
        return False

    print(f"=== STARTING CLINICAL DATABASE SETUP ===")
    print(f"Loading workbook (read-only): {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    
    # Kết nối SQLite
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # -------------------------------------------------------------
    # 1. Parse và xây dựng danh mục bệnh ICD-10 (Bảng Master Sheet)
    # -------------------------------------------------------------
    sheet_name = "ICD10"
    if sheet_name not in wb.sheetnames:
        print(f"[ERROR] Master sheet '{sheet_name}' not found in Excel.")
        conn.close()
        return False
        
    print("Parsing master ICD-10 sheet...")
    sheet = wb[sheet_name]
    records = {}
    row_count = 0
    
    for row in sheet.iter_rows(min_row=4, values_only=True):
        row_count += 1
        if row_count % 3000 == 0:
            print(f"  Processed {row_count} rows...")
            
        if not row or len(row) < 20:
            continue
            
        code = row[16]      # MÃ BỆNH (cột 17)
        name_en = row[18]   # DISEASE NAME (cột 19)
        name_vi = row[19]   # TÊN BỆNH (cột 20)
        
        if code:
            code = str(code).strip().upper()
            name_vi = str(name_vi).strip() if name_vi is not None else ""
            name_en = str(name_en).strip() if name_en is not None else ""
            
            if 3 <= len(code) <= 7:
                records[code] = (code, name_vi, name_en)
                
    print(f"Parsed {len(records)} codes from master sheet.")
    
    # Quét thêm từ các sheet phụ
    for sname in wb.sheetnames:
        if sname == "ICD10" or "Không ghép DRG" in sname or "ko mã bệnh chính" in sname:
            continue
            
        sheet_sub = wb[sname]
        for row in sheet_sub.iter_rows(min_row=4, values_only=True):
            if not row or len(row) < 3:
                continue
            code = row[0]
            name_vi = row[1]
            name_en = row[2]
            
            if code:
                code = str(code).strip().upper()
                if name_vi and str(name_vi).startswith("="): name_vi = ""
                if name_en and str(name_en).startswith("="): name_en = ""
                name_vi = str(name_vi).strip() if name_vi else ""
                name_en = str(name_en).strip() if name_en else ""
                
                if 3 <= len(code) <= 7 and code not in records:
                    records[code] = (code, name_vi, name_en)

    print(f"Total unique ICD-10 codes compiled: {len(records)}")
    
    # Tạo bảng icd10 sạch
    cursor.execute("DROP TABLE IF EXISTS icd10;")
    cursor.execute("""
    CREATE TABLE icd10 (
        code TEXT PRIMARY KEY,
        name_vi TEXT,
        name_en TEXT
    );
    """)
    
    cursor.executemany("INSERT OR REPLACE INTO icd10 VALUES (?, ?, ?);", list(records.values()))
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_icd10_code ON icd10 (code);")
    
    # -------------------------------------------------------------
    # 2. Xây dựng các bảng luật y khoa (Clinical Constraints)
    # -------------------------------------------------------------
    # A. Luật giới tính (Sex constraints)
    print("Building sex constraints rules...")
    cursor.execute("DROP TABLE IF EXISTS icd10_rules_sex;")
    cursor.execute("""
    CREATE TABLE icd10_rules_sex (
        code TEXT PRIMARY KEY,
        allowed_sex TEXT
    );
    """)
    sex_rules = []
    if "A4.1" in wb.sheetnames:
        for row in wb["A4.1"].iter_rows(min_row=4, values_only=True):
            if row and row[0]: sex_rules.append((str(row[0]).strip().upper(), "F"))
    if "A4.2" in wb.sheetnames:
        for row in wb["A4.2"].iter_rows(min_row=4, values_only=True):
            if row and row[0]: sex_rules.append((str(row[0]).strip().upper(), "M"))
    cursor.executemany("INSERT OR REPLACE INTO icd10_rules_sex VALUES (?, ?);", sex_rules)

    # B. Luật độ tuổi (Age constraints)
    print("Building age constraints rules...")
    cursor.execute("DROP TABLE IF EXISTS icd10_rules_age;")
    cursor.execute("""
    CREATE TABLE icd10_rules_age (
        code TEXT PRIMARY KEY,
        min_days INTEGER,
        max_days INTEGER,
        description TEXT
    );
    """)
    age_sheets = {
        "A3.1": (0, 365, "Trẻ sơ sinh (0 - 365 ngày)"),
        "A3.2-A3.3-A3.4": (0, 2*365, "Trẻ nhỏ (0 ngày - 2 tuổi)"),
        "A3.5": (27, 999*365, "Trẻ mới đẻ (trên 27 ngày tuổi)"),
        "A3.6": (365, 999*365, "Trẻ em (1 tuổi trở lên)"),
        "A3.7 - A3.8": (8*365, 19*365, "Tuổi dậy thì (8 – 19 tuổi)"),
        "A3.9": (15*365, 999*365, "Người trưởng thành (trên 15 tuổi)"),
        "A3.10": (30*365, 999*365, "Người lớn (trên 30 tuổi)")
    }
    age_rules = []
    for sname, (min_d, max_d, desc) in age_sheets.items():
        if sname in wb.sheetnames:
            for row in wb[sname].iter_rows(min_row=4, values_only=True):
                if row and row[0]: age_rules.append((str(row[0]).strip().upper(), min_d, max_d, desc))
    cursor.executemany("INSERT OR REPLACE INTO icd10_rules_age VALUES (?, ?, ?, ?);", age_rules)

    # C. Luật mã kép (Dual codes)
    print("Building dual code constraints...")
    cursor.execute("DROP TABLE IF EXISTS icd10_rules_dual;")
    cursor.execute("""
    CREATE TABLE icd10_rules_dual (
        dagger_code TEXT,
        asterisk_code TEXT,
        PRIMARY KEY (dagger_code, asterisk_code)
    );
    """)
    dual_rules = []
    if "A1" in wb.sheetnames:
        for row in wb["A1"].iter_rows(min_row=4, values_only=True):
            if row and len(row) >= 3 and row[0] and row[2]:
                dual_rules.append((str(row[0]).strip().upper(), str(row[2]).strip().upper()))
    cursor.executemany("INSERT OR REPLACE INTO icd10_rules_dual VALUES (?, ?);", dual_rules)

    # D. Luật không được làm bệnh chính (Not primary diagnosis)
    print("Building principal diagnosis constraints...")
    cursor.execute("DROP TABLE IF EXISTS icd10_rules_not_primary;")
    cursor.execute("""
    CREATE TABLE icd10_rules_not_primary (
        code TEXT PRIMARY KEY
    );
    """)
    not_primary_rules = []
    if "A2 Mã ICD10 ko mã bệnh chính" in wb.sheetnames:
        for row in wb["A2 Mã ICD10 ko mã bệnh chính"].iter_rows(min_row=4, values_only=True):
            if row and row[0]: not_primary_rules.append((str(row[0]).strip().upper(),))
    cursor.executemany("INSERT OR REPLACE INTO icd10_rules_not_primary VALUES (?);", not_primary_rules)

    conn.commit()
    conn.close()
    
    # -------------------------------------------------------------
    # 3. Tạo các file context cho LLM
    # -------------------------------------------------------------
    print("Generating LLM flat context files...")
    # JSON phẳng
    dictionary = {code: {"vi": name_vi, "en": name_en} for code, name_vi, name_en in records.values()}
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(dictionary, f, ensure_ascii=False, indent=2)
        
    # Text phẳng
    with open(ICD10_TXT_PATH, "w", encoding="utf-8") as f:
        f.write("=== DANH MỤC MÃ BỆNH ICD-10 SONG NGỮ ===\n\n")
        for code, name_vi, name_en in sorted(list(records.values()), key=lambda x: x[0]):
            f.write(f"[{code}] {name_vi} | {name_en}\n")
            
    # Tạo file RxNorm context trống mặc định
    with open(RXNORM_TXT_PATH, "w", encoding="utf-8") as f:
        f.write("=== DANH MỤC MÃ THUỐC RXNORM ===\n\n")
        
    print(f"=== SETUP COMPLETED SUCCESSFULLY ===")
    print(f"  - Database: {DB_PATH}")
    print(f"  - Dictionary keys: {len(dictionary)}")
    print(f"  - Context ICD-10 file: {ICD10_TXT_PATH}")
    print(f"  - Context RxNorm file: {RXNORM_TXT_PATH}")
    return True

if __name__ == "__main__":
    setup_database()
