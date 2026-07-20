from __future__ import annotations

import argparse
import hashlib
import io
import json
import math
import re
import statistics
import time
import unicodedata
import zipfile
from collections import Counter, defaultdict
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook


RXNCONSO_FIELDS = [
    "RXCUI",
    "LAT",
    "TS",
    "LUI",
    "STT",
    "SUI",
    "ISPREF",
    "RXAUI",
    "SAUI",
    "SCUI",
    "SDUI",
    "SAB",
    "TTY",
    "CODE",
    "STR",
    "SRL",
    "SUPPRESS",
    "CVF",
]

RXNREL_FIELDS = [
    "RXCUI1",
    "AUI1",
    "STYPE1",
    "REL",
    "RXCUI2",
    "AUI2",
    "STYPE2",
    "RELA",
    "RUI",
    "SRUI",
    "SAB",
    "SL",
    "DIR",
    "RG",
    "SUPPRESS",
    "CVF",
]

DEFAULT_RXNORM_TTYS = {
    "IN",
    "PIN",
    "MIN",
    "BN",
    "SCD",
    "SBD",
    "GPCK",
    "BPCK",
    "DF",
    "DFG",
}


def sha256_file(path: Path, block_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while block := stream.read(block_size):
            digest.update(block)
    return digest.hexdigest()


def file_metadata(path: Path) -> dict[str, Any]:
    return {
        "path": path.name,
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def percentile(values: list[int], fraction: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    index = (len(ordered) - 1) * fraction
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return float(ordered[lower])
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (index - lower)


def distribution(values: list[int]) -> dict[str, float | int | None]:
    if not values:
        return {"count": 0, "min": None, "p25": None, "median": None, "p75": None, "p95": None, "max": None, "mean": None}
    return {
        "count": len(values),
        "min": min(values),
        "p25": round(percentile(values, 0.25) or 0, 2),
        "median": round(statistics.median(values), 2),
        "p75": round(percentile(values, 0.75) or 0, 2),
        "p95": round(percentile(values, 0.95) or 0, 2),
        "max": max(values),
        "mean": round(statistics.fmean(values), 2),
    }


def is_safe_zip_member(name: str) -> bool:
    path = PurePosixPath(name)
    return not path.is_absolute() and ".." not in path.parts and not re.match(r"^[A-Za-z]:", name)


def audit_input_zip(path: Path) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        unsafe_members = [info.filename for info in infos if not is_safe_zip_member(info.filename)]
        duplicate_members = [name for name, count in Counter(info.filename for info in infos).items() if count > 1]
        txt_infos = [info for info in infos if not info.is_dir() and info.filename.lower().endswith(".txt")]
        extension_counts = Counter(Path(info.filename).suffix.lower() or "<none>" for info in infos if not info.is_dir())
        root_counts = Counter(PurePosixPath(info.filename).parts[0] for info in infos if PurePosixPath(info.filename).parts)

        byte_lengths: list[int] = []
        char_lengths: list[int] = []
        line_counts: list[int] = []
        utf8_failures: list[dict[str, Any]] = []
        bom_files: list[str] = []
        null_byte_files: list[str] = []
        control_char_files: list[dict[str, Any]] = []
        non_nfc_files: list[str] = []
        empty_files: list[str] = []
        whitespace_only_files: list[str] = []
        hashes: dict[str, list[str]] = defaultdict(list)
        numeric_ids: list[int] = []
        newline_styles = Counter()
        samples: dict[str, str] = {}

        for info in txt_infos:
            raw = archive.read(info)
            byte_lengths.append(len(raw))
            digest = hashlib.sha256(raw).hexdigest()
            hashes[digest].append(info.filename)
            if raw.startswith(b"\xef\xbb\xbf"):
                bom_files.append(info.filename)
            if b"\x00" in raw:
                null_byte_files.append(info.filename)
            if not raw:
                empty_files.append(info.filename)
            try:
                text = raw.decode("utf-8")
            except UnicodeDecodeError as exc:
                utf8_failures.append({"file": info.filename, "error": str(exc)})
                continue
            if not text.strip():
                whitespace_only_files.append(info.filename)
            char_lengths.append(len(text))
            line_counts.append(len(text.splitlines()))
            if "\r\n" in text:
                newline_styles["crlf_files"] += 1
            if "\n" in text.replace("\r\n", ""):
                newline_styles["lf_files"] += 1
            if "\r" in text.replace("\r\n", ""):
                newline_styles["cr_files"] += 1
            controls = sorted({f"U+{ord(ch):04X}" for ch in text if unicodedata.category(ch) == "Cc" and ch not in "\n\r\t"})
            if controls:
                control_char_files.append({"file": info.filename, "characters": controls})
            if unicodedata.normalize("NFC", text) != text:
                non_nfc_files.append(info.filename)
            match = re.search(r"(\d+)\.txt$", info.filename, flags=re.IGNORECASE)
            if match:
                numeric_ids.append(int(match.group(1)))
            if info.filename in {"input/1.txt", "input/31.txt", "input/55.txt", "input/100.txt"}:
                samples[info.filename] = text[:500]

        duplicate_content_groups = [members for members in hashes.values() if len(members) > 1]
        expected_ids = list(range(min(numeric_ids), max(numeric_ids) + 1)) if numeric_ids else []
        missing_ids = sorted(set(expected_ids) - set(numeric_ids))
        duplicate_ids = sorted(value for value, count in Counter(numeric_ids).items() if count > 1)
        bad_crc_member = archive.testzip()

        return {
            "archive": {
                "member_count": len(infos),
                "file_count": sum(not info.is_dir() for info in infos),
                "directory_count": sum(info.is_dir() for info in infos),
                "compressed_bytes": sum(info.compress_size for info in infos),
                "uncompressed_bytes": sum(info.file_size for info in infos),
                "extension_counts": dict(sorted(extension_counts.items())),
                "root_counts": dict(sorted(root_counts.items())),
                "unsafe_members": unsafe_members,
                "duplicate_member_names": duplicate_members,
                "crc_error_member": bad_crc_member,
            },
            "documents": {
                "txt_count": len(txt_infos),
                "numeric_id_min": min(numeric_ids) if numeric_ids else None,
                "numeric_id_max": max(numeric_ids) if numeric_ids else None,
                "missing_numeric_ids": missing_ids,
                "duplicate_numeric_ids": duplicate_ids,
                "all_strict_utf8": not utf8_failures,
                "utf8_failures": utf8_failures,
                "utf8_bom_files": bom_files,
                "empty_files": empty_files,
                "whitespace_only_files": whitespace_only_files,
                "null_byte_files": null_byte_files,
                "unexpected_control_char_files": control_char_files,
                "non_nfc_files": non_nfc_files,
                "newline_styles": dict(newline_styles),
                "byte_length": distribution(byte_lengths),
                "character_length": distribution(char_lengths),
                "line_count": distribution(line_counts),
                "duplicate_content_groups": duplicate_content_groups,
                "samples": samples,
            },
            "annotations": {
                "annotation_files_found": 0,
                "train_split_found": False,
                "validation_split_found": False,
                "ground_truth_schema_available": False,
                "detected_entity_labels": [],
                "detected_assertion_labels": [],
                "detected_relation_labels": [],
            },
        }


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).upper()


def normalized_cell(value: Any) -> str:
    return str(value or "").strip()


def audit_icd10_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet_summary = [
        {"name": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column}
        for sheet in workbook.worksheets
    ]
    expected_sheet = "ICD10"
    if expected_sheet not in workbook.sheetnames:
        return {"sheet_summary": sheet_summary, "expected_sheet_found": False}

    sheet = workbook[expected_sheet]
    header_row = 3
    headers = [normalize_header(cell.value) for cell in sheet[header_row]]
    header_positions = {header: index for index, header in enumerate(headers) if header}
    required_headers = ["MÃ BỆNH", "MÃ BỆNH KHÔNG DẤU", "DISEASE NAME", "TÊN BỆNH"]
    missing_required_headers = [header for header in required_headers if header not in header_positions]

    if missing_required_headers:
        return {
            "sheet_summary": sheet_summary,
            "expected_sheet_found": True,
            "header_row": header_row,
            "headers": headers,
            "missing_required_headers": missing_required_headers,
        }

    idx_code = header_positions["MÃ BỆNH"]
    idx_no_dot = header_positions["MÃ BỆNH KHÔNG DẤU"]
    idx_en = header_positions["DISEASE NAME"]
    idx_vi = header_positions["TÊN BỆNH"]

    record_count = 0
    all_blank_rows = 0
    missing = Counter()
    exact_duplicate_rows = Counter()
    code_counts = Counter()
    no_dot_counts = Counter()
    code_to_names: dict[str, set[tuple[str, str]]] = defaultdict(set)
    no_dot_mismatch_count = 0
    canonical_no_dot_mismatch_count = 0
    no_dot_mismatches: list[dict[str, Any]] = []
    regex_mismatch_count = 0
    canonical_regex_mismatch_count = 0
    regex_mismatches: list[str] = []
    marker_counts = Counter()
    missing_value_samples: list[dict[str, Any]] = []
    first_row_by_record: dict[tuple[str, str, str, str], int] = {}
    exact_duplicate_row_samples: list[dict[str, Any]] = []
    chapter_counts = Counter()
    date_value_types = Counter()
    samples: list[dict[str, Any]] = []
    code_pattern = re.compile(r"^[A-Z][0-9]{2}(?:\.[0-9A-Z]{1,4})?$")

    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = list(row)
        code = normalized_cell(values[idx_code])
        no_dot = normalized_cell(values[idx_no_dot])
        name_en = normalized_cell(values[idx_en])
        name_vi = normalized_cell(values[idx_vi])
        if not any((code, no_dot, name_en, name_vi)):
            all_blank_rows += 1
            continue
        record_count += 1
        record = (code, no_dot, name_en, name_vi)
        exact_duplicate_rows[record] += 1
        if not code:
            missing["code"] += 1
        if not no_dot:
            missing["code_without_dot"] += 1
        if not name_en:
            missing["name_en"] += 1
        if not name_vi:
            missing["name_vi"] += 1
        if (not code or not no_dot or not name_en or not name_vi) and len(missing_value_samples) < 30:
            missing_value_samples.append(
                {
                    "row": row_number,
                    "code": code,
                    "code_without_dot": no_dot,
                    "name_en": name_en,
                    "name_vi": name_vi,
                }
            )
        if record in first_row_by_record and len(exact_duplicate_row_samples) < 30:
            exact_duplicate_row_samples.append(
                {"first_row": first_row_by_record[record], "duplicate_row": row_number, "record": list(record)}
            )
        else:
            first_row_by_record[record] = row_number
        if code:
            code_counts[code] += 1
            code_to_names[code].add((name_en, name_vi))
            chapter_counts[code[0]] += 1
            if "†" in code:
                marker_counts["dagger"] += 1
            if "*" in code:
                marker_counts["asterisk"] += 1
            canonical_code = re.sub(r"[†*]", "", code)
            if not code_pattern.fullmatch(code):
                regex_mismatch_count += 1
                if len(regex_mismatches) < 30:
                    regex_mismatches.append(code)
            if not code_pattern.fullmatch(canonical_code):
                canonical_regex_mismatch_count += 1
        if no_dot:
            no_dot_counts[no_dot] += 1
        if code and no_dot and code.replace(".", "") != no_dot:
            no_dot_mismatch_count += 1
            canonical_code_without_dot = re.sub(r"[†*]", "", code).replace(".", "")
            if canonical_code_without_dot != no_dot:
                canonical_no_dot_mismatch_count += 1
            if len(no_dot_mismatches) < 30:
                no_dot_mismatches.append(
                    {
                        "row": row_number,
                        "code": code,
                        "code_without_dot": no_dot,
                        "canonical_code_without_dot": canonical_code_without_dot,
                    }
                )
        if len(samples) < 5:
            samples.append({"row": row_number, "code": code, "code_without_dot": no_dot, "name_en": name_en, "name_vi": name_vi})
        if len(values) > 21:
            date_value_types[type(values[21]).__name__] += 1

    duplicate_exact_groups = [
        {"record": list(record), "count": count}
        for record, count in exact_duplicate_rows.items()
        if count > 1
    ]
    duplicate_codes = [
        {"code": code, "count": count, "distinct_name_pairs": len(code_to_names[code])}
        for code, count in code_counts.items()
        if count > 1
    ]
    duplicate_no_dot_codes = [
        {"code_without_dot": code, "count": count}
        for code, count in no_dot_counts.items()
        if count > 1
    ]

    return {
        "sheet_summary": sheet_summary,
        "expected_sheet_found": True,
        "header_row": header_row,
        "headers": headers,
        "required_headers": required_headers,
        "missing_required_headers": missing_required_headers,
        "data_start_row": header_row + 1,
        "record_count": record_count,
        "all_blank_rows": all_blank_rows,
        "unique_code_count": len(code_counts),
        "unique_code_without_dot_count": len(no_dot_counts),
        "missing_values": dict(missing),
        "missing_value_samples": missing_value_samples,
        "exact_duplicate_group_count": len(duplicate_exact_groups),
        "exact_duplicate_extra_row_count": sum(group["count"] - 1 for group in duplicate_exact_groups),
        "exact_duplicate_row_samples": exact_duplicate_row_samples,
        "duplicate_code_group_count": len(duplicate_codes),
        "duplicate_code_extra_row_count": sum(group["count"] - 1 for group in duplicate_codes),
        "duplicate_code_samples": duplicate_codes[:30],
        "duplicate_code_without_dot_group_count": len(duplicate_no_dot_codes),
        "duplicate_code_without_dot_extra_row_count": sum(group["count"] - 1 for group in duplicate_no_dot_codes),
        "code_without_dot_mismatch_count": no_dot_mismatch_count,
        "canonical_code_without_dot_mismatch_count": canonical_no_dot_mismatch_count,
        "code_without_dot_mismatch_samples": no_dot_mismatches,
        "code_regex_mismatch_count": regex_mismatch_count,
        "canonical_code_regex_mismatch_count": canonical_regex_mismatch_count,
        "code_regex_mismatch_samples": regex_mismatches,
        "code_marker_counts": dict(marker_counts),
        "chapter_counts": dict(sorted(chapter_counts.items())),
        "date_value_types": dict(date_value_types),
        "samples": samples,
    }


def strip_rrf_trailing_field(parts: list[str]) -> list[str]:
    if parts and parts[-1] == "":
        return parts[:-1]
    return parts


def audit_rxnconso(archive: zipfile.ZipFile, member: str) -> dict[str, Any]:
    field_count_distribution = Counter()
    language_counts = Counter()
    source_counts = Counter()
    tty_counts = Counter()
    suppress_counts = Counter()
    filtered_tty_counts = Counter()
    filtered_rxcuis: set[str] = set()
    filtered_rows = 0
    malformed_samples: list[dict[str, Any]] = []
    filtered_samples: list[dict[str, Any]] = []
    row_count = 0
    decode_error: str | None = None

    try:
        with archive.open(member) as raw_stream:
            with io.TextIOWrapper(raw_stream, encoding="utf-8", errors="strict", newline="") as text_stream:
                for row_count, line in enumerate(text_stream, start=1):
                    parts = strip_rrf_trailing_field(line.rstrip("\r\n").split("|"))
                    field_count_distribution[len(parts)] += 1
                    if len(parts) != len(RXNCONSO_FIELDS):
                        if len(malformed_samples) < 20:
                            malformed_samples.append({"row": row_count, "field_count": len(parts), "preview": parts[:5]})
                        continue
                    record = dict(zip(RXNCONSO_FIELDS, parts))
                    language_counts[record["LAT"]] += 1
                    source_counts[record["SAB"]] += 1
                    tty_counts[record["TTY"]] += 1
                    suppress_counts[record["SUPPRESS"]] += 1
                    if (
                        record["LAT"] == "ENG"
                        and record["SUPPRESS"] == "N"
                        and record["SAB"] == "RXNORM"
                        and record["TTY"] in DEFAULT_RXNORM_TTYS
                    ):
                        filtered_rows += 1
                        filtered_rxcuis.add(record["RXCUI"])
                        filtered_tty_counts[record["TTY"]] += 1
                        if len(filtered_samples) < 12:
                            filtered_samples.append(
                                {
                                    "RXCUI": record["RXCUI"],
                                    "TTY": record["TTY"],
                                    "STR": record["STR"],
                                    "ISPREF": record["ISPREF"],
                                }
                            )
    except UnicodeDecodeError as exc:
        decode_error = str(exc)

    return {
        "member": member,
        "expected_field_count": len(RXNCONSO_FIELDS),
        "row_count": row_count,
        "field_count_distribution": dict(sorted(field_count_distribution.items())),
        "malformed_row_sample_count": len(malformed_samples),
        "malformed_row_samples": malformed_samples,
        "utf8_decode_error": decode_error,
        "language_counts": dict(language_counts.most_common()),
        "source_counts_top_25": dict(source_counts.most_common(25)),
        "tty_counts": dict(tty_counts.most_common()),
        "suppress_counts": dict(suppress_counts.most_common()),
        "default_filter": {
            "LAT": ["ENG"],
            "SAB": ["RXNORM"],
            "SUPPRESS": ["N"],
            "TTY": sorted(DEFAULT_RXNORM_TTYS),
        },
        "filtered_row_count": filtered_rows,
        "filtered_unique_rxcui_count": len(filtered_rxcuis),
        "filtered_tty_counts": dict(filtered_tty_counts.most_common()),
        "filtered_samples": filtered_samples,
    }


def audit_rxnrel(archive: zipfile.ZipFile, member: str) -> dict[str, Any]:
    field_count_distribution = Counter()
    relation_counts = Counter()
    detailed_relation_counts = Counter()
    source_counts = Counter()
    suppress_counts = Counter()
    target_relations = Counter()
    malformed_samples: list[dict[str, Any]] = []
    samples: list[dict[str, Any]] = []
    row_count = 0
    decode_error: str | None = None
    target_names = {"has_ingredient", "tradename_of", "has_dose_form", "consists_of"}

    try:
        with archive.open(member) as raw_stream:
            with io.TextIOWrapper(raw_stream, encoding="utf-8", errors="strict", newline="") as text_stream:
                for row_count, line in enumerate(text_stream, start=1):
                    parts = strip_rrf_trailing_field(line.rstrip("\r\n").split("|"))
                    field_count_distribution[len(parts)] += 1
                    if len(parts) != len(RXNREL_FIELDS):
                        if len(malformed_samples) < 20:
                            malformed_samples.append({"row": row_count, "field_count": len(parts), "preview": parts[:5]})
                        continue
                    record = dict(zip(RXNREL_FIELDS, parts))
                    relation_counts[record["REL"]] += 1
                    detailed_relation_counts[record["RELA"] or "<blank>"] += 1
                    source_counts[record["SAB"]] += 1
                    suppress_counts[record["SUPPRESS"]] += 1
                    if record["RELA"] in target_names:
                        target_relations[record["RELA"]] += 1
                    if len(samples) < 8:
                        samples.append({key: record[key] for key in ("RXCUI1", "REL", "RXCUI2", "RELA", "SAB", "SUPPRESS")})
    except UnicodeDecodeError as exc:
        decode_error = str(exc)

    return {
        "member": member,
        "expected_field_count": len(RXNREL_FIELDS),
        "row_count": row_count,
        "field_count_distribution": dict(sorted(field_count_distribution.items())),
        "malformed_row_sample_count": len(malformed_samples),
        "malformed_row_samples": malformed_samples,
        "utf8_decode_error": decode_error,
        "rel_counts": dict(relation_counts.most_common()),
        "rela_counts_top_40": dict(detailed_relation_counts.most_common(40)),
        "source_counts_top_25": dict(source_counts.most_common(25)),
        "suppress_counts": dict(suppress_counts.most_common()),
        "requested_relation_counts": dict(target_relations.most_common()),
        "samples": samples,
    }


def audit_rxnorm_zip(path: Path, scan_relations: bool) -> dict[str, Any]:
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        members = {info.filename: info for info in infos}
        required = ["rrf/RXNCONSO.RRF", "rrf/RXNREL.RRF"]
        readme_name = next((name for name in members if name.startswith("Readme_Full_") and name.endswith(".txt")), None)
        readme_text = archive.read(readme_name).decode("utf-8", errors="replace") if readme_name else ""
        result: dict[str, Any] = {
            "archive": {
                "member_count": len(infos),
                "compressed_bytes": sum(info.compress_size for info in infos),
                "uncompressed_bytes": sum(info.file_size for info in infos),
                "unsafe_members": [info.filename for info in infos if not is_safe_zip_member(info.filename)],
                "required_members": {name: name in members for name in required},
                "required_member_sizes": {
                    name: {"compressed_bytes": members[name].compress_size, "uncompressed_bytes": members[name].file_size}
                    for name in required
                    if name in members
                },
                "prescribe_variant_present": "prescribe/rrf/RXNCONSO.RRF" in members,
            },
            "readme": {
                "member": readme_name,
                "preview": readme_text[:2000],
            },
        }
        if "rrf/RXNCONSO.RRF" in members:
            result["rxnconso"] = audit_rxnconso(archive, "rrf/RXNCONSO.RRF")
        if scan_relations and "rrf/RXNREL.RRF" in members:
            result["rxnrel"] = audit_rxnrel(archive, "rrf/RXNREL.RRF")
        else:
            result["rxnrel"] = {"scan_skipped": True, "reason": "--skip-rxnrel was used"}
        return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Phase 1 data audit for the Clinical NLP lab.")
    parser.add_argument("--root", type=Path, default=Path.cwd())
    parser.add_argument("--output", type=Path, default=Path("reports/phase1_audit.json"))
    parser.add_argument("--skip-rxnrel", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root = args.root.resolve()
    output = args.output if args.output.is_absolute() else root / args.output
    input_zip = root / "input.zip"
    icd10_xlsx = root / "ICD10.xlsx"
    rxnorm_zip = root / "RxNorm_full_07062026.zip"

    started = time.time()
    report: dict[str, Any] = {
        "audit_version": "1.0.0",
        "root": str(root),
        "files": {},
        "input_zip": {},
        "icd10": {},
        "rxnorm": {},
        "phase1_constraints": {
            "training_authorized": False,
            "training_performed": False,
            "private_test_used_for_fitting": False,
        },
    }

    for path in (input_zip, icd10_xlsx, rxnorm_zip):
        if not path.exists():
            report["files"][path.name] = {"exists": False}
        else:
            report["files"][path.name] = {"exists": True, **file_metadata(path)}

    if input_zip.exists():
        print("Auditing input.zip...", flush=True)
        report["input_zip"] = audit_input_zip(input_zip)
    if icd10_xlsx.exists():
        print("Auditing ICD10.xlsx...", flush=True)
        report["icd10"] = audit_icd10_workbook(icd10_xlsx)
    if rxnorm_zip.exists():
        print("Auditing RxNorm zip (streaming; this can take a while)...", flush=True)
        report["rxnorm"] = audit_rxnorm_zip(rxnorm_zip, scan_relations=not args.skip_rxnrel)

    report["elapsed_seconds"] = round(time.time() - started, 2)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="\n") as stream:
        json.dump(report, stream, ensure_ascii=False, indent=2, sort_keys=False)
        stream.write("\n")
    print(f"Wrote {output}", flush=True)


if __name__ == "__main__":
    main()
